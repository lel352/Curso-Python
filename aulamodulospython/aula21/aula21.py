"""
https://openpyxl.readthedocs.io/en/stable/
pip install openpyxl
pipenv install openpyxl
"""
import openpyxl
from random import uniform


pedidos = openpyxl.load_workbook('pedidos.xlsx')
nome_planilhas = pedidos.sheetnames
planilha1 = pedidos['Página1']

print(planilha1['B4'].value)  # retorna um objeto da Celula
print('*'*25)

for campo in planilha1['b']: # planilha1['B'] retorna um tupla com toda a coluna B
    print(campo.value)

print('*'*25)
for linha in planilha1['a1:c2']:  #pegando um range de valores de a1 a c2    pega linha
    for coluna in linha:  # coluna da linha
        print(coluna.value)

print('*'*25)
for linha in planilha1:  # for em toda a planilha
    for coluna in linha:
        print(coluna.value)


print('*'*25)
for linha in planilha1:  # for em toda a planilha
    for i in range(len(linha)):
        if linha[i].value is not None:
            print(linha[i].value, end=' ')
    print('')

print('*' * 25)

planilha1['B3'].value = 2200
pedidos.save('nova_planilha.xlsx') # alterou o valor do B3


print('*' * 25)
for linha in range(5, 16):
    numero_pedido = linha - 1
    planilha1.cell(linha, 1).value = numero_pedido
    planilha1.cell(linha, 2).value = 1200 + linha

    preco = round(uniform(10, 100), 2)
    planilha1.cell(linha, 3).value = preco

pedidos.save('nova_planilha.xlsx')



planilha = openpyxl.Workbook()
planilha.create_sheet('Planilha1', 0)
planilha.create_sheet('Planilha2', 1)

planilha1 = planilha['Planilha1']
planilha2 = planilha['Planilha2']

for linha in range(1, 11):
    numero_pedido = linha - 1
    planilha1.cell(linha, 1).value = numero_pedido
    planilha1.cell(linha, 2).value = 1200 + linha

    preco = round(uniform(10, 100), 2)
    planilha2.cell(linha, 3).value = preco

for linha in range(1, 11):
    planilha2.cell(linha, 1).value = f'Luiz {linha} {round(uniform(10, 100), 2)}'
    planilha2.cell(linha, 2).value = f'Otávio {linha} {round(uniform(10, 100), 2)}'
    planilha2.cell(linha, 3).value = f'Joãozinho {linha} {round(uniform(10, 100), 2)}'

planilha.save('nova_planilha2.xlsx')
